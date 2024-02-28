from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.action_chains import ActionChains
from msedge.selenium_tools import Edge, EdgeOptions
import sys
import getpass
import time
import os
import re
import fnmatch
import openpyxl
import ctypes
import warnings
import shutil
import ctypes
import tkinter as tk
from tkinter import simpledialog
from tkinter import messagebox

   
# using getlogin() returning username
UserName = os.getlogin()

# Initialize variables
CFSaas = ""
ProjectNumber = ""
ProjectNumber = ""
ClientName = ""
ProjectName = ""
ProjectCode = ""
Mainfolderpath = ""
RulesetP = ""
IncASCII = ""
UserEmail = ""
EmailEntered = ""
templateid = ""
Path = ""

#Folder Define
BeforeUserName = 'C:/Users/'
AfterUserName = '/Sermo/RealTime Pro Projects - Documents/'
Tempfolder = 'Template - RTP projects'
Tempfolder2 = 'Template - RTP projects2'
Mainfolderpath = BeforeUserName + UserName + AfterUserName


def get_project_details():
    global CFSaas, ProjectNumber, ClientName, ProjectName, ProjectCode, Mainfolderpath, RulesetP, IncASCII, UserEmail, EmailEntered, templateid, Path

    def on_closing():
        if tk.messagebox.askokcancel("Quit", "Do you want to quit?"):
            root.destroy()
            sys.exit()
    
    root = tk.Tk()
    root.title("Project Details")
    root.protocol("WM_DELETE_WINDOW", on_closing)
    font = ("Calibri", 11)
    entry_width = 35
    width4DD = 29
    
     # Create a Frame to hold the form
    form_frame = tk.Frame(root, padx=80, pady=40)
    form_frame.pack()

    # Project Type
    label_type = tk.Label(form_frame, text="Project Platform:")
    label_type.grid(row=0, column=0, padx=(0, 70))
    project_type_options = ["Forsta", "Confirmit"]
    ProjectType = tk.StringVar(root)
    ProjectType.set(project_type_options[0])  # Set the default value
    type_menu = tk.OptionMenu(form_frame, ProjectType, *project_type_options)
    type_menu.grid(row=0, column=1, sticky="w")
    type_menu.config(width=width4DD) 


    # Project Number
    label_number = tk.Label(form_frame, text="Project Number:")
    entry_number = tk.Entry(form_frame, width=entry_width)
    label_number.grid(row=2, column=0, sticky="w")
    entry_number.grid(row=2, column=1)

    # Client Name
    label_client = tk.Label(form_frame, text="Client Name:")
    entry_client = tk.Entry(form_frame, width=entry_width)
    label_client.grid(row=4, column=0, sticky="w")
    entry_client.grid(row=4, column=1)

    # Project Name
    label_name = tk.Label(form_frame, text="Project Name:")
    entry_name = tk.Entry(form_frame, width=entry_width)
    label_name.grid(row=6, column=0, sticky="w")
    entry_name.grid(row=6, column=1)

    # Project Code
    label_code = tk.Label(form_frame, text="Forsta/Confirmit pcode:")
    entry_code = tk.Entry(form_frame, width=entry_width)
    label_code.grid(row=8, column=0, sticky="w")
    entry_code.grid(row=8, column=1)

    # Project Survey Template ID
    label_template = tk.Label(form_frame, text="Survey Template ID:")
    entry_template = tk.Entry(form_frame, width=entry_width)
    label_template.grid(row=10, column=0, sticky="w")
    entry_template.grid(row=10, column=1)

    # Path
    PathEntered = "Y"
    label_Path = tk.Label(form_frame, text="DM Folder Path:")
    entry_Path = tk.Entry(form_frame, width=entry_width)
    label_Path.grid(row=12, column=0, sticky="w")
    entry_Path.grid(row=12, column=1) 

   # Create Ruleset label and drop-down
    label_ruleset = tk.Label(form_frame, text="Create Ruleset:")
    label_ruleset.grid(row=14, column=0, sticky="w")
    ruleset_options = ["Yes", "No"]
    CreateRuleset = tk.StringVar(root)
    CreateRuleset.set(ruleset_options[0])
    ruleset_menu = tk.OptionMenu(form_frame, CreateRuleset, *ruleset_options)
    ruleset_menu.grid(row=14, column=1, sticky="w")
    ruleset_menu.config(width=width4DD)

    # Include ASCII label and drop-down
    label_ascii = tk.Label(form_frame, text="Include ASCII in Ruleset:")
    label_ascii.grid(row=16, column=0, sticky="w")
    ascii_options = ["Yes", "No"]
    IncludeASCII = tk.StringVar(root)
    IncludeASCII.set(ascii_options[1])
    ascii_menu = tk.OptionMenu(form_frame, IncludeASCII, *ascii_options)
    ascii_menu.grid(row=16, column=1, sticky="w")
    ascii_menu.config(width=width4DD)

    # Email input
    if os.path.exists(Mainfolderpath):
        Mainfolderpath = Mainfolderpath
        UserEmail = UserName + '@sermo.com'
    else:
        EmailEntered = "Y"
        Mainfolderpath = os.getcwd().replace('\\','/') + '/'
        label_Email = tk.Label(form_frame, text="Your email id:")
        entry_Email = tk.Entry(form_frame, width=entry_width)
        label_Email.grid(row=18, column=0, sticky="w")
        entry_Email.grid(row=18, column=1)
   
    # Button frame
    button_frame = tk.Frame(form_frame)
    button_frame.grid(row=20, columnspan=2, pady=10)


    # Button to submit
    def validate_and_submit():
        global CFSaas, ProjectNumber, ClientName, ProjectName, ProjectCode, Mainfolderpath, RulesetP, IncASCII, UserEmail, EmailEntered, templateid, Path
        
        CFSaas = ProjectType.get()
        ProjectNumber = entry_number.get()
        ClientName = entry_client.get()
        ProjectName = entry_name.get()
        ProjectCode = entry_code.get()
        templateid = entry_template.get()
        Path = entry_Path.get()
        RulesetP = CreateRuleset.get()
        IncASCII = IncludeASCII.get()
        if EmailEntered == "Y":
            UserEmail = entry_Email.get()

        if not ProjectNumber or not ClientName or not ProjectName or not UserEmail:
            messagebox.showerror("Validation Error", "Project Number, Client Name, and Project Name are required.")
             
        else:
            #print(f"Project Type: {CFSaas}")
            #print(f"Project Number: {ProjectNumber}")
            #print(f"Client Name: {ClientName}")
            #print(f"Project Name: {ProjectName}")
            #print(f"Project Code: {ProjectCode}")
            #print(f"Survey Template ID: {templateid}")
            #print(f"Ruleset: {RulesetP}")
            #print(f"Include ASCII: {IncASCII}")
            #print(f"User Email: {UserEmail}")

            root.destroy()
        
    submit_button = tk.Button(button_frame, text="Submit", command=validate_and_submit, bg="#90EE90")
    submit_button.grid(row=20, column=0, padx=(0, 50), pady=(25, 0))

    # Cancel button
    def cancel():
        root.destroy()
        sys.exit()  

    cancel_button = tk.Button(button_frame, text="Cancel", command=cancel, bg="#FFA07A")
    cancel_button.grid(row=20, column=1, pady=(25, 0))

    root.mainloop()

if __name__ == "__main__":
    get_project_details()

if ProjectNumber == "" or ClientName == "" or ProjectName == "":
    print("All mandatory (*) field should be filled")
    ctypes.windll.user32.MessageBoxW(0, "All mandatory (*) field should be filled", "Unscussful", 16)
    sys.exit()

# Checking if the folder exist
if os.path.exists(Path):
    source_folder = Path
    print(Path)
else:
    print(Path)
    print("Folder not found")
    ctypes.windll.user32.MessageBoxW(0, "Folder not found, please input correct DM folder path", "Folder not found", 16)
    sys.exit()


#Conversion to already existing codes
if CFSaas == "Forsta":
    CFSaas = "y"
else:
    CFSaas = "n"

if RulesetP == "Yes":
    RulesetP = "y"
else:
    RulesetP = "n"

if IncASCII == "Yes":
    IncASCII = "y"
else:
    IncASCII = "n"    



    
# get the full path of the file you want to update
file_path = Path + '/'
fileName = '!Delivery_manager.xlsm'

# Checking if DM exist
if os.path.exists(file_path + fileName):
    file_path = Path + '/'
    #print(Path)
else:
    print("!Delivery_manager.xlsm not found")
    ctypes.windll.user32.MessageBoxW(0, "!Delivery_manager.xlsm file not found, copy 07 Survey Data Operations folder from TEMPLATE_FOLDER", "DM file not found", 16)
    sys.exit

# Load the Excel file
workbook = openpyxl.load_workbook(file_path+fileName , data_only=False, keep_vba=True)
sh = workbook["Parameters"]
# Select the cell to fill with text
cellB2 = sh['B2'] 
cellB3 = sh['B3']
cellB4 = sh['B4']
cellB7 = sh['B7']
cellB8 = sh['B8']
cellB9 = sh['B9']
cellB16 = sh['B16']
cellB17 = sh['B17']

ncode = int(ProjectNumber)
cellB2.value = ProjectNumber
cellB3.value = ProjectName
cellB4.value = ClientName

if CFSaas == 'n' or CFSaas == 'N':
    cellB7.value = '<ProjectCode>_SurveyData.xlsx'
    cellB8.value = '<ProjectCode>_SurveyData.sav'
    cellB9.value = '<ProjectCode>_SurveyData_responseid.asc'
    cellB16.value = 'ccode=Country'
    cellB17.value = 'responseid, respid, pin, pass'
    sh2 = workbook["Operations"]
    # Select the cell to fill with text
    sh2['C2'] = 'TRUE'
    sh2['C3'] = 'FALSE'
    sh2['D2'] = 'TRUE'
    sh2['D3'] = 'FALSE'
    sh2['E2'] = 'TRUE'
    sh2['E3'] = 'FALSE'
    sh2['F2'] = 'TRUE'
    sh2['F3'] = 'FALSE'

#Save the Excel macro files
workbook.save(file_path+fileName)

# get the new name of the file
new_name = file_path + '!' + ProjectNumber +' Delivery_manager.xlsm'

# rename the file
os.rename(file_path+fileName, new_name)

# get the full path of the file you want to rename and change the name
file_path2 = Path + '/DataMapLayout/Autoexports/'
fileName2 = '[enter project number] Autoexport.xlsx'

# Checking if folder exist
if os.path.exists(file_path2 + fileName2) and os.path.exists(Path + '/DataMapLayout/Rules/'):
    file_path2 = Path + '/DataMapLayout/Autoexports/'
else:
    print("Autoexports / Rules folder not found")
    ctypes.windll.user32.MessageBoxW(0, "Autoexports / Rules folder not found", "Folder not found", 16)
    sys.exit

workbook = openpyxl.load_workbook(file_path2+fileName2)
sh = workbook["Autoexport"]
# Select the cell to fill with text
cellA2 = sh['A2']
cellA3 = sh['A3']
cellA4 = sh['A4']
cellA5 = sh['A5']
cellA6 = sh['A6']
cellA7 = sh['A7']
cellF2 = sh['F2']
cellF3 = sh['F3']
cellF4 = sh['F4']
cellF5 = sh['F5']
cellF6 = sh['F6']
cellF7 = sh['F7']
cellI2 = sh['I2']
cellI3 = sh['I3']
cellI4 = sh['I4']
cellI5 = sh['I5']
cellI6 = sh['I6']
cellI7 = sh['I7']

cellA2.value = ProjectCode
cellA3.value = ProjectCode
cellA4.value = ProjectCode
cellA5.value = ProjectCode
cellA6.value = ProjectCode
cellA7.value = ProjectCode

cellF2.value = templateid
cellF3.value = templateid
cellF4.value = templateid
cellF5.value = templateid
cellF6.value = templateid
cellF7.value = templateid
if CFSaas == 'n' or CFSaas == 'N':
    cellI2.value = 'False'
    cellI3.value = 'False'
    cellI4.value = 'False'
    cellI5.value = 'False'
    cellI6.value = 'False'
    cellI7.value = 'False'
else:
    cellI2.value = 'True'
    cellI3.value = 'True'
    cellI4.value = 'True'
    cellI5.value = 'True'
    cellI6.value = 'True'
    cellI7.value = 'True'
workbook.save(file_path2+fileName2)

new_name2 = file_path2 + ProjectNumber +'_Autoexport.xlsx'
os.rename(file_path2+fileName2, new_name2)

# ------- Rule Creator---------------

if ProjectCode == "" or templateid == "":
    ctypes.windll.user32.MessageBoxW(0, "Folder created, rules not created as all the required information not provided", "Successful", 0)
    print("Folder created, Rules not created as all the required information not provided")
    sys.exit()
else:
    import xml.etree.ElementTree as ET
    from datetime import datetime
    

    # create the file structure
    def create_rule(pcode, project_code, template_id, frmt, my_email, my_part, forsta):
        panelRule = ET.Element('PanelRule')
        panelRule.set('xmlns:xsd', "http://www.w3.org/2001/XMLSchema")
        panelRule.set('xmlns:xsi', "http://www.w3.org/2001/XMLSchema-instance")

        panelId = ET.SubElement(panelRule, 'PanelId')
        panelId.text = "-1"
        created = ET.SubElement(panelRule, 'Created')
        created.text = "2016-10-21T09:36:13.343"
        lastUpdated = ET.SubElement(panelRule, 'LastUpdated')
        lastUpdated.text = "2016-10-21T09:42:51.837"
        createdBy = ET.SubElement(panelRule, 'CreatedBy')
        createdBy.text = "SERMO"
        lastUpdated = ET.SubElement(panelRule, 'LastUpdated')
        lastUpdated.text = "SERMO"
        propertyValues = ET.SubElement(panelRule, 'PropertyValues')
        ownerUserName = ET.SubElement(panelRule, 'OwnerUserName')
        ruleId = ET.SubElement(panelRule, 'RuleId')
        ruleId.text = "111111"
        ruleName = ET.SubElement(panelRule, 'RuleName')
        ruleName.text = "{}_{}".format(pcode,frmt)
        isTemporary = ET.SubElement(panelRule, 'IsTemporary')
        isTemporary.text='false'
        lastExecutedBy = ET.SubElement(panelRule, 'LastExecutedBy')
        lastExecuted = ET.SubElement(panelRule, 'LastExecuted')
        lastExecuted.text = '0001-01-01T00:00:00'
        companyId = ET.SubElement(panelRule, 'CompanyId')
        companyId.text = "2"
        status = ET.SubElement(panelRule, 'Status')
        status.text = 'Enabled'

        conditionExpression = ET.SubElement(panelRule, 'ConditionExpression')
        if forsta==True:
            conditionExpression.text = 'isTest = "0" AND NOT ISNULL(rdout) AND NOT IN(respondentStatus, \"DUPLICATE\", \"RESET\", \"REMOVE\", \"RECO\", \"UNVERIFIED\")'
        elif forsta==False:
            conditionExpression.text = 'xtest = "0" AND (VRFD = "1" OR ISNULL(VRFD)) AND NOT ISNULL(rdout) AND NOT IN(respondentStatus, \"DUPLICATE\", \"RESET\", \"REMOVE\", \"RECO\", \"UNVERIFIED\")'
            
        variables = ET.SubElement(panelRule, 'Variables')
        action = ET.SubElement(panelRule, 'Action')
        loopActions = ET.SubElement(panelRule, 'LoopActions')
        LoopIds = ET.SubElement(panelRule, "LoopIds")
        globalScript = ET.SubElement(panelRule, "GlobalScript")
        globalProperties = ET.SubElement(panelRule, "GlobalProperties")
        postScript = ET.SubElement(panelRule, "PostScript")
        comment = ET.SubElement(panelRule, "Comment")
        #comment.text = frmt
        if forsta==True:
            comment.text = "{}_{}_SurveyData{}_sFTP".format(pcode,frmt,my_part)

        if forsta==False:
            comment.text = "{}_{}_SurveyData{}".format(pcode,frmt,my_part)

        selectedCount = ET.SubElement(panelRule, "SelectedCount")
        selectedCount.text = "0"
        qualifiedCount = ET.SubElement(panelRule, "QualifiedCount")
        qualifiedCount.text = "0"
        updatedCount = ET.SubElement(panelRule, "UpdatedCount")
        updatedCount.text = "0"
        currentRuleTaskId = ET.SubElement(panelRule, "CurrentRuleTaskId")
        currentRuleTaskId.text = "-1"
        isAdHoc = ET.SubElement(panelRule, "IsAdHoc")
        isAdHoc.text = "false"
        
        targetSettings = ET.SubElement(panelRule, "TargetSettings")
        if frmt == 'Excel':
            targetSettings.set('xsi:type', "ExcelTargetSettings")
        elif frmt == 'SPSS':
            targetSettings.set('xsi:type', "SavTargetSettings")
        elif frmt == 'ASCII':
            targetSettings.set('xsi:type', "SssDataTargetSettings")
            
        mappedFields = ET.SubElement(targetSettings, "MappedFields")
        fileName = ET.SubElement(targetSettings, "FileName")
        #fileName.text = "{}_SurveyData{}".format(pcode,my_part)

        #2022-11-22: Igoris added
        if forsta==True:
            if frmt == 'Excel':
                fileName.text = "{}_SurveyData{}_excel".format(pcode,my_part)
            elif frmt == 'SPSS':
                fileName.text = "{}_SurveyData{}_spss".format(pcode,my_part)
            elif frmt == 'ASCII':
                fileName.text = "{}_SurveyData{}_ascii".format(pcode,my_part)
        if forsta==False:
            fileName.text = "{}_SurveyData{}".format(pcode,my_part)

        #2022-11-22: Igoris added
        if forsta==True:
            Uncompressed = ET.SubElement(targetSettings, "Uncompressed")
            Uncompressed.text = "true"

        emailOptions = ET.SubElement(targetSettings, "EmailOptions")
        replyTo = ET.SubElement(emailOptions, "ReplyTo")
        subject = ET.SubElement(emailOptions, "Subject")
        encodingCodePage = ET.SubElement(emailOptions, "EncodingCodePage")
        encodingCodePage.set('xsi:nil',"true")
        HtmlBody = ET.SubElement(emailOptions, 'HtmlBody')
        PlainTextBody = ET.SubElement(emailOptions, 'PlainTextBody')

        emailAddress = ET.SubElement(targetSettings, "EmailAddress")
        emailAddress.text = my_email
        overrideFileName = ET.SubElement(targetSettings, "OverrideFileName")
        overrideFileName.text = "true"
        fileTransferType = ET.SubElement(targetSettings, "FileTransferType")
        if forsta==True:
            #fileTransferType.text = "Email"
            fileTransferType.text = "ExternalFtpServer"
        elif forsta==False:
            fileTransferType.text = "FtpServer"
        encryptFile = ET.SubElement(targetSettings, "EncryptFile")
        encryptFile.text = "false"
        useInternally = ET.SubElement(targetSettings, "UseInternally")
        useInternally.text = "false"
        openTextWidth = ET.SubElement(targetSettings, "OpenTextWidth")
        if frmt=="Excel" or frmt=='ASCII':
            openTextWidth.text = "-1"
        elif frmt=='SPSS':
            openTextWidth.text = "200"
        loopHandling = ET.SubElement(targetSettings, "LoopHandling")
        loopHandling.text = "SeparateFiles"
        loopPosition = ET.SubElement(targetSettings, "LoopPosition")
        loopPosition.text = "AsQuestionnaire"
        if forsta==True:
            externalFtpTargetParameters = ET.SubElement(targetSettings, "ExternalFtpTargetParameters")
            externalFtpTargetParameters.set("FolderName", "Download")
            externalFtpTargetParameters.set("HostName", "share.sermo.com")
            externalFtpTargetParameters.set("UserName", "w7o3MYw9ECcoD7ZWAUUbadApgYQvH2NR/lHshG5lrdE=")
            externalFtpTargetParameters.set("Password", "5uojQn2LKas9OJQhBmw1hw==")
            externalFtpTargetParameters.set("UseSsh", "true")
            externalFtpTargetParameters.set("HostFingerprint", "c7:e1:28:09:2a:b3:e5:df:23:70:dc:2e:aa:bd:c9:74")
            externalFtpTargetParameters.set("AlwaysTrustThisHost", "false")
        elif forsta==False:
            externalFtpTargetParameters = ET.SubElement(targetSettings, "ExternalFtpTargetParameters")
            externalFtpTargetParameters.set("AlwaysTrustThisHost", "false")
            externalFtpTargetParameters.set("HostFingerprint", "")
            externalFtpTargetParameters.set("UseSsh", "false")
            externalFtpTargetParameters.set("Password", "CCGYoC3KgCySSEmApXHqOA==")
            externalFtpTargetParameters.set("UserName", "4BxDU5hon8xwv2+CdvRLAQ==")
            externalFtpTargetParameters.set("HostName", "")
            externalFtpTargetParameters.set("FolderName", "")
        
        
        #recodeMultis = ET.SubElement(targetSettings, "RecodeMultis")
        #recodeMultis.text = "false" !!not in excel xml

        if frmt=='Excel':
            excelVersion = ET.SubElement(targetSettings, "ExcelVersion")
            excelVersion.text = "MsExcel2007"
        if frmt=='ASCII':
            SssTemplateId = ET.SubElement(targetSettings, "SssTemplateId")
            SssTemplateId.text = template_id
            IncludeSchema = ET.SubElement(targetSettings, "IncludeSchema")
            IncludeSchema.text = "true"
            IncludeData = ET.SubElement(targetSettings, "IncludeData")
            IncludeData.text = "true"
            CodePage = ET.SubElement(targetSettings, "CodePage")
            CodePage.text = "28591"
            CodePageForSchema = ET.SubElement(targetSettings, "CodePageForSchema")
            CodePageForSchema.text = "65001"
            IncludeConfirmitMetaTags = ET.SubElement(targetSettings, "IncludeConfirmitMetaTags")
            IncludeConfirmitMetaTags.text = "false"
            
        sourceSettings = ET.SubElement(panelRule, "SourceSettings")
        sourceSettings.set("xsi:type", "SurveyDataSourceSettings")
        projectIds = ET.SubElement(sourceSettings, "ProjectIds")
        string_projectcode = ET.SubElement(projectIds, "string")
        string_projectcode.text = project_code

        databaseType = ET.SubElement(sourceSettings, "DatabaseType")
        databaseType.text = "Production"
        ruleDateFilterType = ET.SubElement(sourceSettings, "RuleDateFilterType")
        ruleDateFilterType.text = "None"
        isIncrementalUpdate = ET.SubElement(sourceSettings, "IsIncrementalUpdate")
        isIncrementalUpdate.text = "false"
        lastChangeTrackingVersion = ET.SubElement(sourceSettings, "LastChangeTrackingVersion")
        lastChangeTrackingVersion.text = "0"
        keywordFilter = ET.SubElement(sourceSettings, "KeywordFilter")
        rowLimit = ET.SubElement(sourceSettings, "RowLimit")
        rowLimit.text = "0"
        responseFilter = ET.SubElement(sourceSettings, "ResponseFilter")
        responseStatus = ET.SubElement(responseFilter, "ResponseStatus")       
        responseStatus.text = "Complete"
        filterTemplateId = ET.SubElement(sourceSettings, "FilterTemplateId")
        exportFieldLabelSourceType = ET.SubElement(sourceSettings, "ExportFieldLabelSourceType")
        
        if frmt=='Excel':
            filterTemplateId.text = template_id
            exportFieldLabelSourceType.text = "Project"
        elif frmt=='SPSS':
            filterTemplateId.text = template_id
            exportFieldLabelSourceType.text = "Template"
        elif frmt=='ASCII':
            filterTemplateId.text = "0"
            exportFieldLabelSourceType.text = "Project"

        hideTemplate = ET.SubElement(sourceSettings, "HideTemplate")
        hideTemplate.text = "false"
        allowFilterVarsNotExistInDb = ET.SubElement(sourceSettings, "AllowFilterVarsNotExistInDb")
        allowFilterVarsNotExistInDb.text = "false"
        showVarNotExistWarning = ET.SubElement(sourceSettings, "ShowVarNotExistWarning")
        showVarNotExistWarning.text = "false"
        allowFilterAnswersNotExistInDb = ET.SubElement(sourceSettings, "AllowFilterAnswersNotExistInDb")
        allowFilterAnswersNotExistInDb.text = "false"
        showAnswerNotExistWarning = ET.SubElement(sourceSettings, "ShowAnswerNotExistWarning")
        showAnswerNotExistWarning.text = "false"
        AddLabels = ET.SubElement(sourceSettings, "AddLabels")
        AddLabels.text = "false"
        labelLanguage = ET.SubElement(sourceSettings, "LabelLanguage")
        labelLanguage.text = "9"
        labelType = ET.SubElement(sourceSettings, "LabelType")
        labelType.text = "QuestionId"
        questionElementDescriptionType = ET.SubElement(sourceSettings, "QuestionElementDescriptionType")
        questionElementDescriptionType.text = "AnswerQuestionLabel"
        openEndHandling = ET.SubElement(sourceSettings, "OpenEndHandling")
        openEndHandling.text = "IncludeOpenEnds"
        
        source = ET.SubElement(panelRule, "Source")
        source_typeId = ET.SubElement(source, "TypeId")
        source_typeId.text = "SurveyData"
        source_text = ET.SubElement(source, "Text")
        source_text.text = "Survey Database"

        
        target = ET.SubElement(panelRule, "Target")
        target_typeId = ET.SubElement(target, "TypeId")
        target_text = ET.SubElement(target, "Text")
        if frmt=='Excel':
            target_typeId.text = "Excel"
            target_text.text = "Excel File"
        elif frmt=='SPSS':
            target_typeId.text = "SpssSav"
            target_text.text = "SPSS File (sav)"
        elif frmt=='ASCII':
            target_typeId.text = "SssDataFile"
            target_text.text = "Triple-S XML (Standard)"
            
        ruleType = ET.SubElement(panelRule, "RuleType")
        ruleType.text = "Normal"
        
        sTVS = ET.SubElement(panelRule, "SourceAndTargetValidationSettings")
        allowVarInSourceNotInTarget = ET.SubElement(sTVS, "AllowVarInSourceNotInTarget")
        allowVarInSourceNotInTarget.text = "false"
        showVarInSourceNotInTargetWarning = ET.SubElement(sTVS, "ShowVarInSourceNotInTargetWarning")
        showVarInSourceNotInTargetWarning.text = "false"
        allowVarInTargetNotInSource = ET.SubElement(sTVS, "AllowVarInTargetNotInSource")
        allowVarInTargetNotInSource.text = "false"
        showVarInTargetNotInSourceWarning = ET.SubElement(sTVS, "ShowVarInTargetNotInSourceWarning")
        showVarInTargetNotInSourceWarning.text = "false"
        allowAnswerInSourceNotInTarget = ET.SubElement(sTVS, "AllowAnswerInSourceNotInTarget")
        allowAnswerInSourceNotInTarget.text = "false"
        showAnswerInSourceNotInTargetWarning = ET.SubElement(sTVS, "ShowAnswerInSourceNotInTargetWarning")
        showAnswerInSourceNotInTargetWarning.text = "false"
        allowAnswerInTargetNotInSource = ET.SubElement(sTVS, "AllowAnswerInTargetNotInSource")
        allowAnswerInTargetNotInSource.text = "false"
        showAnswerInTargetNotInSourceWarning = ET.SubElement(sTVS, "ShowAnswerInTargetNotInSourceWarning")
        showAnswerInTargetNotInSourceWarning.text = "false"

        dCPD = ET.SubElement(panelRule, "DataCentralProjectDescription")
        dCPD.set('DesktopVersionOnly', "false")
        dCPD.set('Version', "1")
        dCPD.set('Id', "00000000-0000-0000-0000-000000000000")

        inputSurveys = ET.SubElement(dCPD, "InputSurveys")
        outputSurveys = ET.SubElement(dCPD, "OutputSurveys")
        messages_surveys = ET.SubElement(dCPD, "Messages")
        logs_surveys = ET.SubElement(dCPD, "Logs")
        reports_surveys = ET.SubElement(dCPD, "Reports")

        dataCentralProjectId = ET.SubElement(panelRule, "DataCentralProjectId")
        dataCentralProjectId.text = "0"
        dataCentralProjectLastUpdated = ET.SubElement(panelRule, "DataCentralProjectLastUpdated")
        dataCentralProjectLastUpdated.text = "0001-01-01T00:00:00"

        # create a new XML file with the results
        mydata = ET.tostring(panelRule)
        if forsta == True:
            myfile = open("{}_{}_SurveyData{}_sFTP.xml".format(pcode,frmt,my_part), "wb")
        if forsta == False:
            myfile = open("{}_{}_SurveyData{}.xml".format(pcode,frmt,my_part), "wb")
        myfile.write(mydata)




    import os
    import sys
    os.chdir(Path + '/DataMapLayout/Rules')

    #from xml_structure_script import create_rule

    if CFSaas == 'n' or CFSaas == 'N':
        forsta = False
    else:
        forsta = True
    project_number = ProjectNumber 
    confirmit_pcode = ProjectCode
    part_of_study = ""            #Specify Sufix for multiple sets (for exaple if you have HCP and PATS, write _HCP or _PATS - one run at a time)
    template_id = templateid
    ASCII = True
    my_email = UserEmail

    for f in ['Excel', 'SPSS']:
        create_rule(pcode = project_number, project_code = confirmit_pcode, template_id = template_id, frmt=f, my_email=my_email, my_part = part_of_study, forsta=forsta)
        if ASCII==True:
            create_rule(pcode = project_number, project_code = confirmit_pcode, template_id = template_id, frmt='ASCII', my_email=my_email, my_part = part_of_study, forsta=forsta)

    rulesecreated = "Y"






    
    # ------- Auto Export Creator---------------
    
    import xml.etree.ElementTree as ET
    from datetime import datetime
    from openpyxl import load_workbook
    import os
    import sys

    def create_rule(xml_loc, pcode, project_code, template_id, sermo_email, mail_to_deliver, venid, doc_type, comm, forsta):
        panelRule = ET.Element('PanelRule')
        panelRule.set('xmlns:xsd', "http://www.w3.org/2001/XMLSchema")
        panelRule.set('xmlns:xsi', "http://www.w3.org/2001/XMLSchema-instance")

        panelId = ET.SubElement(panelRule, 'PanelId')
        panelId.text = "-1"
        created = ET.SubElement(panelRule, 'Created')
        created.text = "2016-10-21T09:36:13.343"
        lastUpdated = ET.SubElement(panelRule, 'LastUpdated')
        lastUpdated.text = "2016-10-21T09:42:51.837"
        createdBy = ET.SubElement(panelRule, 'CreatedBy')
        createdBy.text = "SERMO"
        lastUpdated = ET.SubElement(panelRule, 'LastUpdated')
        lastUpdated.text = "SERMO"
        propertyValues = ET.SubElement(panelRule, 'PropertyValues')
        ownerUserName = ET.SubElement(panelRule, 'OwnerUserName')
        ruleId = ET.SubElement(panelRule, 'RuleId')
        ruleId.text = "111111"
        ruleName = ET.SubElement(panelRule, 'RuleName')
        ruleName.text = "{}_data".format(pcode)
        isTemporary = ET.SubElement(panelRule, 'IsTemporary')
        isTemporary.text='false'
        lastExecutedBy = ET.SubElement(panelRule, 'LastExecutedBy')
        lastExecuted = ET.SubElement(panelRule, 'LastExecuted')
        lastExecuted.text = '0001-01-01T00:00:00'
        companyId = ET.SubElement(panelRule, 'CompanyId')
        companyId.text = "2"
        status = ET.SubElement(panelRule, 'Status')
        status.text = 'Enabled'
        conditionExpression = ET.SubElement(panelRule, 'ConditionExpression')
        if venid=='Client':
            if forsta=='True':
                conditionExpression.text = 'isTest = "0" AND NOT IN(respondentStatus, \"DUPLICATE\", \"RESET\", \"REMOVE\", \"RECO\", \"UNVERIFIED\")'
            elif forsta=='False':
                conditionExpression.text = '(VRFD = "1" OR ISNULL(VRFD)) AND xtest = "0" AND NOT IN(respondentStatus, \"DUPLICATE\", \"RESET\", \"REMOVE\", \"RECO\", \"UNVERIFIED\")'
        else:
            if forsta=='True':
                    conditionExpression.text = 'vendorId = {}'.format(venid)
            elif forsta=='False':
                    conditionExpression.text = 'IN(VENID, "{}")'.format(venid)
        variables = ET.SubElement(panelRule, 'Variables')
        action = ET.SubElement(panelRule, 'Action')
        loopActions = ET.SubElement(panelRule, 'LoopActions')
        loopIds = ET.SubElement(panelRule, "loopIds")
        globalScript = ET.SubElement(panelRule, "GlobalScript")
        globalProperties = ET.SubElement(panelRule, "GlobalProperties")
        postScript = ET.SubElement(panelRule, "PostScript")
        comment = ET.SubElement(panelRule, "Comment")
        if venid=='Client':
            if comm==None:
                 comment.text = "{} AutoExport Client ({})".format(pcode,doc_type)
            else:
                comment.text = "{} AutoExport Client ({}) - {}".format(pcode,doc_type,comm)
        else:
            if comm==None:
                 comment.text = "{} AutoExport Venid={} ({})".format(pcode, venid, doc_type)
            else:
                comment.text = "{} AutoExport Venid={} ({}) - {}".format(pcode, venid, doc_type,comm)

        selectedCount = ET.SubElement(panelRule, "SelectedCount")
        selectedCount.text = "0"
        qualifiedCount = ET.SubElement(panelRule, "QualifiedCount")
        qualifiedCount.text = "0"
        updatedCount = ET.SubElement(panelRule, "UpdatedCount")
        updatedCount.text = "0"
        currentRuleTaskId = ET.SubElement(panelRule, "CurrentRuleTaskId")
        currentRuleTaskId.text = "-1"
        isAdHoc = ET.SubElement(panelRule, "IsAdHoc")
        isAdHoc.text = "false"
        targetSettings = ET.SubElement(panelRule, "TargetSettings")

        if doc_type=='SpssSav':
            targetSettings.set('xsi:type', "SavTargetSettings")
        if (doc_type=='Excel') or (doc_type=='ExcelWithLabels'):
            targetSettings.set('xsi:type', "ExcelTargetSettings")
        mappedFields = ET.SubElement(targetSettings, "MappedFields")
        fileName = ET.SubElement(targetSettings, "FileName")
        if comm!=None:
            fileName.text = "{}_AutoExport({})".format(pcode,comm)
        else:
            fileName.text = "{}_AutoExport".format(pcode) 
        emailOptions = ET.SubElement(targetSettings, "EmailOptions")
        replyTo = ET.SubElement(emailOptions, "ReplyTo")
        replyTo.text = sermo_email
        subject = ET.SubElement(emailOptions, "Subject")
        if comm!=None:
            subject.text = "{} - {} Data Auto Export".format(pcode, comm)
        else:
            subject.text = "{} - Data Auto Export".format(pcode)
        encodingCodePage = ET.SubElement(emailOptions, "EncodingCodePage")
        encodingCodePage.text = "65001" #this is UTF8, for ANSI it's 28591
        htmlBody = ET.SubElement(emailOptions, "HtmlBody")
        if venid=='Client':
            htmlBody.text = "<p><span style='font-family:\"Calibri\",\"serif\";color:blue;font-size:16'>Hi,<br><br>Please find attached the automatic data export for subject study.<br></span></p> <p><span style='font-family:\"Calibri\",\"serif\";color:blue;font-size:14'><i>*Please note that the data you are receiving contains respondents that have not yet been processed through our respondent data check. We ask that you reserve judgement on all data until our team has had a chance to fully review and provide you with their comments/findings and, if needed, an updated file. If, at any point, you noticed an issue that we have not addressed, we ask that you please bring them to our attention. Also note that the number of respondents provided in the file may not match our reporting tool, as we are constantly gaining additional completes. Please keep in mind that there might be some errors due to respondent behavior on the link as well.</i><br></span></p><p><span style='font-family:\"Calibri\",\"serif\";color:#018da9;font-size:13'><br>Sincerely,<br><b>Data Processing Team</b></span></p> <A href='http://SERMO.com/'><IMG border='0' hspace='0' alt='www.SERMO.com' align='baseline' src='https://s3.amazonaws.com/emailsiginternal.sermo.com/logo-color.png' NOSEND='1' width='158' height='48'></A>".format(pcode)
        else:
            htmlBody.text = "<p><span style='font-family:\"Calibri\",\"serif\";color:blue;font-size:16'>Hi,<br><br>Please find attached the automatic data export for subject study.<br></span></p> <p><span style='font-family:\"Calibri\",\"serif\";color:#018da9;font-size:13'><br>Sincerely,<br><b>Data Processing Team</b></span></p><A href='http://SERMO.com/'><IMG border='0' hspace='0' alt='www.SERMO.com' align='baseline' src='https://s3.amazonaws.com/emailsiginternal.sermo.com/logo-color.png' NOSEND='1' width='158' height='48'></A>".format(pcode)
        plainTextBody = ET.SubElement(emailOptions, "PlainTextBody")
        plainTextBody.text = "   "

        emailAddress = ET.SubElement(targetSettings, "EmailAddress")
        emailAddress.text = mail_to_deliver
        overrideFileName = ET.SubElement(targetSettings, "OverrideFileName")
        overrideFileName.text = "true"
        fileTransferType = ET.SubElement(targetSettings, "FileTransferType")
        fileTransferType.text = "Email"
        encryptFile = ET.SubElement(targetSettings, "EncryptFile")
        encryptFile.text = "false"
        useInternally = ET.SubElement(targetSettings, "UseInternally")
        useInternally.text = "false"
        openTextWidth = ET.SubElement(targetSettings, "OpenTextWidth")
        openTextWidth.text = "3500"
        loopHandling = ET.SubElement(targetSettings, "LoopHandling")
        loopHandling.text = "SingleFile"
        loopPosition = ET.SubElement(targetSettings, "LoopPosition")
        loopPosition.text = "AsQuestionnaire"
        externalFtpTargetParameters = ET.SubElement(targetSettings, "ExternalFtpTargetParameters")
        externalFtpTargetParameters.set("AlwaysTrustThisHost", "false")
        externalFtpTargetParameters.set("HostFingerprint", "")
        externalFtpTargetParameters.set("UseSsh", "false")
        externalFtpTargetParameters.set("Password", "")
        externalFtpTargetParameters.set("UserName", "")
        externalFtpTargetParameters.set("HostName", "")
        externalFtpTargetParameters.set("FolderName", "")
        recodeMultis = ET.SubElement(targetSettings, "RecodeMultis")
        recodeMultis.text = "false"

        excelVersion = ET.SubElement(targetSettings, "ExcelVersion")
        excelVersion.text = "MsExcel2007"

        sourceSettings = ET.SubElement(panelRule, "SourceSettings")
        sourceSettings.set("xsi:type", "SurveyDataSourceSettings")
        projectIds = ET.SubElement(sourceSettings, "ProjectIds")
        string_projectcode = ET.SubElement(projectIds, "string")
        string_projectcode.text = project_code

        databaseType = ET.SubElement(sourceSettings, "DatabaseType")
        databaseType.text = "Production"
        ruleDateFilterType = ET.SubElement(sourceSettings, "RuleDateFilterType")
        ruleDateFilterType.text = "None"
        isIncrementalUpdate = ET.SubElement(sourceSettings, "IsIncrementalUpdate")
        isIncrementalUpdate.text = "false"
        lastChangeTrackingVersion = ET.SubElement(sourceSettings, "LastChangeTrackingVersion")
        lastChangeTrackingVersion.text = "0"
        keywordFilter = ET.SubElement(sourceSettings, "KeywordFilter")
        rowLimit = ET.SubElement(sourceSettings, "RowLimit")
        rowLimit.text = "0"
        responseFilter = ET.SubElement(sourceSettings, "ResponseFilter")
        if venid=='Client':
            responseStatus = ET.SubElement(responseFilter, "ResponseStatus")
            responseStatus.text = "Complete"

        filterTemplateId = ET.SubElement(sourceSettings, "FilterTemplateId")
        filterTemplateId.text = template_id

        exportFieldLabelSourceType = ET.SubElement(sourceSettings, "ExportFieldLabelSourceType")
        if venid=='Client':
            if doc_type=='Excel':
                exportFieldLabelSourceType.text = "Project"
            else:
                exportFieldLabelSourceType.text = "Template"
        else:
            exportFieldLabelSourceType.text = "Template"


        hideTemplate = ET.SubElement(sourceSettings, "HideTemplate")
        hideTemplate.text = "false"
        allowFilterVarsNotExistInDb = ET.SubElement(sourceSettings, "AllowFilterVarsNotExistInDb")
        allowFilterVarsNotExistInDb.text = "false"
        showVarNotExistWarning = ET.SubElement(sourceSettings, "ShowVarNotExistWarning")
        showVarNotExistWarning.text = "false"
        allowFilterAnswersNotExistInDb = ET.SubElement(sourceSettings, "AllowFilterAnswersNotExistInDb")
        allowFilterAnswersNotExistInDb.text = "false"
        showAnswerNotExistWarning = ET.SubElement(sourceSettings, "ShowAnswerNotExistWarning")
        showAnswerNotExistWarning.text = "false"
        AddLabels = ET.SubElement(sourceSettings, "AddLabels")
        AddLabels.text = "false"
        labelLanguage = ET.SubElement(sourceSettings, "LabelLanguage")
        labelLanguage.text = "9"
        
        labelType = ET.SubElement(sourceSettings, "LabelType")
        if venid=='Client':
            labelType.text = "QuestionId"
        else:
            labelType.text = "TitleAndText"

        questionElementDescriptionType = ET.SubElement(sourceSettings, "QuestionElementDescriptionType")
        questionElementDescriptionType.text = "AnswerQuestionLabel"
        openEndHandling = ET.SubElement(sourceSettings, "OpenEndHandling")
        openEndHandling.text = "IncludeOpenEnds"


        source = ET.SubElement(panelRule, "Source")
        source_typeId = ET.SubElement(source, "TypeId")
        source_typeId.text = "SurveyData"
        source_text = ET.SubElement(source, "Text")
        source_text.text = "Survey Database"

        target = ET.SubElement(panelRule, "Target")
        target_typeId = ET.SubElement(target, "TypeId")
        target_text = ET.SubElement(target, "Text")

        target_typeId.text = doc_type
        if doc_type=='Excel':
            target_text.text = "Excel File"
        elif doc_type=='ExcelWithLabels':
            target_text.text = "Excel File (Answer Codes as Labels)"
        elif doc_type=='SpssSav':
            target_text.text = "SPSS File (sav)"

        ruleType = ET.SubElement(panelRule, "RuleType")
        ruleType.text = "Normal"
        sTVS = ET.SubElement(panelRule, "SourceAndTargetValidationSettings")

        allowVarInSourceNotInTarget = ET.SubElement(sTVS, "AllowVarInSourceNotInTarget")
        allowVarInSourceNotInTarget.text = "false"
        showVarInSourceNotInTargetWarning = ET.SubElement(sTVS, "ShowVarInSourceNotInTargetWarning")
        showVarInSourceNotInTargetWarning.text = "false"
        allowVarInTargetNotInSource = ET.SubElement(sTVS, "AllowVarInTargetNotInSource")
        allowVarInTargetNotInSource.text = "false"
        showVarInTargetNotInSourceWarning = ET.SubElement(sTVS, "ShowVarInTargetNotInSourceWarning")
        showVarInTargetNotInSourceWarning.text = "false"
        allowAnswerInSourceNotInTarget = ET.SubElement(sTVS, "AllowAnswerInSourceNotInTarget")
        allowAnswerInSourceNotInTarget.text = "false"
        showAnswerInSourceNotInTargetWarning = ET.SubElement(sTVS, "ShowAnswerInSourceNotInTargetWarning")
        showAnswerInSourceNotInTargetWarning.text = "false"
        allowAnswerInTargetNotInSource = ET.SubElement(sTVS, "AllowAnswerInTargetNotInSource")
        allowAnswerInTargetNotInSource.text = "false"
        showAnswerInTargetNotInSourceWarning = ET.SubElement(sTVS, "ShowAnswerInTargetNotInSourceWarning")
        showAnswerInTargetNotInSourceWarning.text = "false"

        dCPD = ET.SubElement(panelRule, "DataCentralProjectDescription")
        dCPD.set('DesktopVersionOnly', "false")
        dCPD.set('Version', "1")
        dCPD.set('Id', "00000000-0000-0000-0000-000000000000")

        inputSurveys = ET.SubElement(dCPD, "InputSurveys")
        outputSurveys = ET.SubElement(dCPD, "OutputSurveys")
        messages_surveys = ET.SubElement(dCPD, "Messages")
        logs_surveys = ET.SubElement(dCPD, "Logs")
        reports_surveys = ET.SubElement(dCPD, "Reports")

        dataCentralProjectId = ET.SubElement(panelRule, "DataCentralProjectId")
        dataCentralProjectId.text = "0"
        dataCentralProjectLastUpdated = ET.SubElement(panelRule, "DataCentralProjectLastUpdated")
        dataCentralProjectLastUpdated.text = "0001-01-01T00:00:00"

        # create a new XML file with the results
        mydata = ET.tostring(panelRule)
        if venid=='Client':
            if comm==None:
                myfile = open(xml_loc+"{}_AutoExport_Client_({}).xml".format(pcode,doc_type), "wb")
            else:
                myfile = open(xml_loc+"{}_AutoExport_Client_({})_{}.xml".format(pcode,doc_type,comm), "wb")
        else:
            if comm==None:
                myfile = open(xml_loc+"{}_AutoExport_Venid={}_({}).xml".format(pcode, venid, doc_type), "wb")
            else:
                myfile = open(xml_loc+"{}_AutoExport_Venid={}_({})_{}.xml".format(pcode, venid, doc_type,comm), "wb")
        myfile.write(mydata)




    if os.getcwd() not in sys.path:
        sys.path.append(os.getcwd())  


    import tkinter
    from tkinter import filedialog
    tkinter.Tk().withdraw()
    #exe_file = filedialog.askopenfilename(filetypes=[("Excel file","*.xlsx"),("Excel file", "*.xls")])  
    #if exe_file[-10:]!='xport.xlsx' and exe_file[-9:]!='xport.xls':
    #   input('Please make sure you select an AutoExport excel file')
    #    sys.exit()
            


    columns = ['ConfirmitPcode', 'Owner', 'Venids', 'RecipientEmails', 'ReplyToEmail', 'FilterTemplateId', 'TargetTypeID', 'Suffix', 'Forsta']

    #workbook = load_workbook(filename=exe_file)
    workbook = openpyxl.load_workbook(file_path2+ProjectNumber +'_Autoexport.xlsx')
    sheet = workbook.active
        
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        if  row!= tuple(columns):
            input('Please make sure that columns are named as follows: '+ str(columns))
            sys.exit()


    xml_creation_location = file_path2
    pcode = ProjectNumber
    project_code = {}
    vendor_ids = {}
    recipient_emails = {}
    sermo_email = {}
    template_id = {}
    doc_type = {}
    comm = {}
    forsta = {}

    invalid_cells=[]


    for n,row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        if (row[0]!=None) & (row[1]=='Vendor') & (None not in [i for i in row[2:-2]]) & (row[-1]!=None):
            project_code[n] = row[0]
            vendor_ids[n] = row[2]
            recipient_emails[n] = row[3]
            sermo_email[n] = row[4]
            template_id[n] = str(row[5])
            doc_type[n] = row[6]
            comm[n]= row[7]
            forsta[n]=row[8]
        elif (row[0]!=None) & (row[1]=='Client') & (None not in [i for i in row[3:-2]]) & (row[-1]!=None):
            project_code[n] = row[0]
            vendor_ids[n] = 'Client'
            recipient_emails[n] = row[3]
            sermo_email[n] = row[4]
            template_id[n] = str(row[5])
            doc_type[n] = row[6]
            comm[n]= row[7]
            forsta[n]=row[8]
        else:
            if row[1]=='Vendor' or row[1]==None:
                invalid_columns = [i for i,e in enumerate(row[:-1]) if e==None]
                invalid_cells.append([n+2, invalid_columns])
            else:
                invalid_columns = [i for i,e in enumerate(row[:-1]) if e==None]
                if 2 in invalid_columns:
                    invalid_columns.remove(2)
                invalid_cells.append([n+2, invalid_columns])
                
    if len(invalid_cells)>0:
        print('File is invalid.')
        for i in invalid_cells:
            input('prompt: Please check row: '+str(i[0])+', column(s): '+str([c for c in columns if columns.index(c) in i[1]]))
        sys.exit()   



    for n in range(len(project_code)):
        create_rule(xml_loc = xml_creation_location, pcode = pcode, project_code = project_code[n],
                    template_id = template_id[n], sermo_email = sermo_email[n],
                    mail_to_deliver = recipient_emails[n], venid = vendor_ids[n], 
                    doc_type=doc_type[n], comm=comm[n], forsta=forsta[n])




    #input('prompt: Files created.')



##----------------------------------------------------Importing rulese and creating ruleset------------------------------------------------------------



if (RulesetP == 'y' or RulesetP == 'Y') and (CFSaas == 'y' or CFSaas == 'Y') and rulesecreated == "Y":
    pcode = str(ProjectNumber)
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    #Get driver and its options and services
    edge_options = EdgeOptions()
    edge_options.use_chromium = True
    edge_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    edge_options.add_argument("--headless")
    driver = Edge(executable_path='C:/Python3/msedgedriver.exe', options=edge_options)
    #driver = Edge(executable_path='C:/Python3/msedgedriver.exe')

    driver.get("https://author.euro.confirmit.com/confirm/authoring/Confirmit.aspx")

    wait = WebDriverWait(driver, 10)
    username_input = wait.until(EC.visibility_of_element_located((By.ID, "username")))

    Username = os.getlogin()
    
    def login_window():
        global Username, Password

        def on_closing():
            if tk.messagebox.askokcancel("Quit", "Do you want to quit?"):
                root.destroy()
                driver.close()
                driver.quit()
                print("Rules XML created - rules not imported and ruleset not created")
                ctypes.windll.user32.MessageBoxW(0, "Folder and Rules XML created - rules not imported and ruleset not created", "Successful", 0)
                sys.exit()
        
        root = tk.Tk()
        root.title("Login")
        root.geometry("300x150")
        root.protocol("WM_DELETE_WINDOW", on_closing)
        # Create a Frame to hold the form
        form_frame = tk.Frame(root, padx=20, pady=20)
        form_frame.grid(row=0, column=0)

        label_username = tk.Label(form_frame, text="Username:")
        label_username.grid(row=0, column=0, padx=(0, 40))
        entry_username = tk.Entry(form_frame)
        entry_username.grid(row=0, column=1, sticky="w")
        entry_username.insert(0, Username)

        label_password = tk.Label(form_frame, text="Password:")
        label_password.grid(row=1, column=0, sticky="w")
        entry_password = tk.Entry(form_frame, show="*")
        entry_password.grid(row=1, column=1, sticky="w")

        def validate_login():
            global Username, Password

            Username = entry_username.get()
            Password = entry_password.get()

            if validate_password(Username, Password):
                
                root.destroy()
                

                #Click DataProcessing button    
                dbutton = driver.find_element(By.ID, "__button_dataprocessing_inner")
                dbutton.click()
                RulePath = Path + '/DataMapLayout/Rules/'

                if os.path.exists(source_folder):
                    #print("Rules folder found in the directory")
                    xlsxfile = RulePath + pcode + "_Excel_SurveyData_sFTP.xml"
                    SPSSfile = RulePath + pcode + "_SPSS_SurveyData_sFTP.xml"
                    ASCIIfile = RulePath + pcode + "_ASCII_SurveyData_sFTP.xml"
                else:
                    print("Rule folder not found in the directory, folder and rules created, rules not imported")
                    ctypes.windll.user32.MessageBoxW(0, "Rule folder not found in the directory, folder and rules created, rules not imported", "Rules Folder not found", 16)
                    driver.close()
                    driver.quit()
                    sys.exit()
                

                # List of file paths to be uploaded
                file_paths = [xlsxfile, SPSSfile, ASCIIfile]
                print("Importing Rules")
                for filepath in file_paths:
                    uploaded_file_name = os.path.splitext(os.path.basename(filepath))[0]
                    
                    #Clicking Import button 
                    wait = WebDriverWait(driver, 5)
                    username_input = wait.until(EC.visibility_of_element_located((By.ID, "__button_dp_rulelist")))

                    rslbutton = driver.find_element(By.ID, "__button_dp_rulelist")
                    rslbutton.click()

                    iframe = driver.find_element(By.ID, "main_frame")
                    driver.switch_to.frame(iframe)

                    wait = WebDriverWait(driver, 10)
                    username_input = wait.until(EC.visibility_of_element_located((By.ID, "ctl03_miImport")))

                    rslbutton = driver.find_element(By.ID, "ctl03_miImport")
                    rslbutton.click()

                    driver.switch_to.default_content()

                    #Uploading the XML file
                    try:
                        wait = WebDriverWait(driver, 5)
                        iframemain = driver.find_element(By.ID, "main_frame")
                        driver.switch_to.frame(iframemain)
                        wait = WebDriverWait(driver, 5)
                        username_input = wait.until(EC.visibility_of_element_located((By.ID, "jobFile")))
                        file_input = driver.find_element(By.ID, "jobFile")
                        file_input.send_keys(filepath)
                        text_box = driver.find_element(By.ID, "panelruleName")
                        text_box.clear()
                        text_box.send_keys(uploaded_file_name)
                        ok1 = driver.find_element(By.ID, "okbutton")
                        ok1.click()
                        driver.switch_to.default_content() 
                    except Exception as e:
                        print("Rules created, rules not imported, browser error")
                        ctypes.windll.user32.MessageBoxW(0, "Folder and rules created, rules not imported, browser error", "Browser error", 16)
                        driver.switch_to.default_content()
                        driver.close()
                        driver.quit()
                        sys.exit()

                #Creating ruleset
                print("Creating ruleset, please wait..")
                wait = WebDriverWait(driver, 5)
                username_input = wait.until(EC.visibility_of_element_located((By.ID, "__button_dp_rulesetlist")))

                rslbutton = driver.find_element(By.ID, "__button_dp_rulesetlist")
                rslbutton.click()

                iframem2 = driver.find_element(By.ID, "main_frame")
                driver.switch_to.frame(iframem2)

                newruleset = driver.find_element(By.ID, "ctl02_miNew")
                newruleset.click()
                driver.switch_to.default_content()

                try:
                    if IncASCII == "y" or IncASCII == "Y":
                        RulesetName = pcode + "_Excel_SPSS_ASCII_SurveyData_sFTP"
                    else:
                        RulesetName = pcode + "_Excel_SPSS_SurveyData_sFTP"

                    wait = WebDriverWait(driver, 5)
                    iframemain = driver.find_element(By.ID, "main_frame")
                    driver.switch_to.frame(iframemain)

                    text_box = driver.find_element(By.ID, "textRuleSetName")
                    text_box.clear()
                    text_box.send_keys(RulesetName)
                    ok2 = driver.find_element(By.ID, "buttonMenu_buttonOk")
                    ok2.click()
                    driver.switch_to.default_content()

                    wait = WebDriverWait(driver, 5)
                    iframemain = driver.find_element(By.ID, "main_frame")
                    driver.switch_to.frame(iframemain)
                    rulesetid = driver.find_element(By.ID, "ucGeneral_txtID")
                    rulesetid_text = rulesetid.get_attribute("value")
                except Exception as e:
                    print("Ruleset name should be unique, delete existing ruleset with name "+RulesetName)
                    ctypes.windll.user32.MessageBoxW(0, "Ruleset name should be unique,", "Unique ruleset name error", 16)
                    driver.close()
                    driver.quit()
                    sys.exit()

                element = driver.find_element(By.XPATH, '//*[@id="tsRuleSetEditor"]/tbody/tr[2]/td[4]')
                element.click()

                driver.switch_to.default_content()
                wait = WebDriverWait(driver, 5)
                iframemain = driver.find_element(By.ID, "main_frame")
                driver.switch_to.frame(iframemain)

                time.sleep(3)
                rulename1 = pcode + "_Excel_SurveyData_sFTP"
                text_box2 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miGen_txtRuleName")
                text_box2.clear()

                actions = ActionChains(driver)

                for char in rulename1:
                    text_box2.send_keys(char)
                    time.sleep(0.1)
                actions.perform()

                option_xpath = f'//div[@class="yui-ac-bd"]/ul/li[contains(text(), "{rulename1}")]'
                option_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
                option_element.click()

                ok3 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miAdd")
                ok3.click()

                time.sleep(1)
                rulename2 = pcode + "_SPSS_SurveyData_sFTP"
                text_box2 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miGen_txtRuleName")
                text_box2.clear()

                actions = ActionChains(driver)

                for char in rulename2:
                    text_box2.send_keys(char)
                    time.sleep(0.1)
                actions.perform()

                option_xpath = f'//div[@class="yui-ac-bd"]/ul/li[contains(text(), "{rulename2}")]'
                option_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
                option_element.click()

                ok3 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miAdd")
                ok3.click()

                if IncASCII == "y" or IncASCII == "Y":
                    time.sleep(3)
                    rulename2 = pcode + "_ASCII_SurveyData_sFTP"
                    text_box2 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miGen_txtRuleName")
                    text_box2.clear()

                    actions = ActionChains(driver)

                    for char in rulename2:
                        text_box2.send_keys(char)
                        time.sleep(0.1)
                    actions.perform()

                    option_xpath = f'//div[@class="yui-ac-bd"]/ul/li[contains(text(), "{rulename2}")]'
                    option_element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
                    option_element.click()

                    ok3 = driver.find_element(By.ID, "ucManagement_ruleManagementMenu_miAdd")
                    ok3.click()
                 
                driver.close()
                driver.quit()
                
                # Load the Excel file
                workbook = openpyxl.load_workbook(new_name, data_only=False, keep_vba=True)
                sh = workbook["Parameters"]
                # Select the cell to fill with text
                cellB6 = sh['B6'] 
                cellB6.value = rulesetid_text
                workbook.save(new_name)
                print("Rules imported and ruleset created, DM updated")
                ctypes.windll.user32.MessageBoxW(0, "Rules imported and ruleset created, DM updated", "Successful", 0)
                sys.exit()

    
            else:
                messagebox.showerror("Validation Error", "Invalid username or password.")
                entry_password.delete(0, 'end')
                

        submit_button = tk.Button(form_frame, text="Submit", command=validate_login, bg="#90EE90")
        submit_button.grid(row=3, column=0, padx=(0, 30), pady=(25, 0))

        def cancel():
            root.destroy()
            driver.close()
            driver.quit()
            print("Folder and Rules XML created - rules not imported and ruleset not created")
            ctypes.windll.user32.MessageBoxW(0, "Folder and Rules XML created - rules not imported and ruleset not created", "Successful", 0)
            sys.exit()  

        cancel_button = tk.Button(form_frame, text="Cancel", command=cancel, bg="#FFA07A")
        cancel_button.grid(row=3, column=1, pady=(25, 0))

        root.mainloop()
    def validate_password(username, password):
        try:
            username_input = driver.find_element(By.ID, "username")
            if username_input.get_attribute("value"):
                username_input.clear()
            username_input.send_keys(username)
            password_input = driver.find_element(By.ID, "password")
            if password_input.get_attribute("value"):
                password_input.clear()
            password_input.send_keys(password)
        except:
            print("Not able to pass the userid and password")

        try:
            login_button = driver.find_element(By.ID, "btnlogin")
            login_button.click()
            print("Validating username and password")
            wait = WebDriverWait(driver, 5)
            username_input = wait.until(EC.visibility_of_element_located((By.ID, "__button_dataprocessing_inner")))
            return True
        except:
            return False
            

    if __name__ == "__main__":
        login_window()

else:
    print("Rules XML created - rules not imported and ruleset not created")
    ctypes.windll.user32.MessageBoxW(0, "Folder and Rules XML created - rules not imported and ruleset not created", "Successful", 0)
    sys.exit()






